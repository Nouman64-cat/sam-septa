import io
import os
import threading
from contextlib import asynccontextmanager
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import FastAPI, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from sqlmodel import Session, select

from db import create_db_and_tables, engine
from models import ScrapeJob


# ── In-memory job registry ────────────────────────────────────────────────────
# Mirrors the DB ScrapeJob row for fast status polling without DB hits.

_jobs: dict[str, dict] = {}


def _new_job_id(prefix: str) -> str:
    return f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"


def _finish_job(job_id: str, status: str, record_count: int, error: str | None = None):
    """Update both the in-memory registry and the DB ScrapeJob row."""
    now = datetime.utcnow()
    _jobs[job_id]["status"]       = status
    _jobs[job_id]["record_count"] = record_count
    _jobs[job_id]["error"]        = error
    with Session(engine) as s:
        job = s.exec(select(ScrapeJob).where(ScrapeJob.job_id == job_id)).first()
        if job:
            job.status        = status
            job.finished_at   = now
            job.record_count  = record_count
            job.error_message = error
            s.add(job)
            s.commit()


# ── Excel helpers ─────────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _styled_workbook(sheet_name: str, headers: list[str], rows: list[list]) -> io.BytesIO:
    """Build a styled openpyxl workbook and return it as an in-memory BytesIO stream."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header row
    ws.append(headers)
    for cell in ws[1]:
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[1].height = 30

    # Data rows
    for row in rows:
        ws.append(row)

    # Auto-fit column widths (capped at 60)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


# ── App setup ─────────────────────────────────────────────────────────────────

@asynccontextmanager
async def lifespan(app: FastAPI):
    create_db_and_tables()
    yield


app = FastAPI(lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "http://localhost:3001",
        "http://127.0.0.1:3001",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── Include route modules ────────────────────────────────────────────────────

from routes.sam import router as sam_router
from routes.septa import router as septa_router
from routes.unison import router as unison_router
from routes.dibbs import router as dibbs_router
from routes.naics import router as naics_router
from routes.eval_config import router as eval_config_router

app.include_router(sam_router)
app.include_router(septa_router)
app.include_router(unison_router)
app.include_router(dibbs_router)
app.include_router(naics_router)
app.include_router(eval_config_router)


# ── Misc ──────────────────────────────────────────────────────────────────────

@app.get("/")
async def home():
    return {"message": "SAM-SEPTA Scraper API"}


# ── Job status / stop ─────────────────────────────────────────────────────────

@app.get("/status/{job_id}")
async def get_job_status(job_id: str):
    if job_id not in _jobs:
        return JSONResponse({"error": "Unknown job ID"}, status_code=404)
    job = _jobs[job_id]
    return JSONResponse({
        "status":       job["status"],        # running | done | stopped | error
        "record_count": job["record_count"],  # live count updated per bid
        "error":        job["error"],
    })


@app.post("/stop/{job_id}")
async def stop_job(job_id: str):
    if job_id not in _jobs:
        return JSONResponse({"error": "Unknown job ID"}, status_code=404)
    job = _jobs[job_id]
    if job["status"] != "running":
        return JSONResponse({"success": False, "message": "Job is not currently running."})
    job["stop_event"].set()
    return JSONResponse({
        "success": True,
        "message": "Stop signal sent - finishing current bid then saving.",
    })


# ── Scrape jobs list ──────────────────────────────────────────────────────────

@app.get("/jobs")
async def list_jobs(scraper: Optional[str] = Query(default=None)):
    """List all scrape jobs, newest first. Filter by scraper='sam' or 'septa'."""
    with Session(engine) as s:
        query = select(ScrapeJob).order_by(ScrapeJob.started_at.desc())
        if scraper:
            query = query.where(ScrapeJob.scraper == scraper)
        jobs = s.exec(query).all()

    return JSONResponse([
        {
            "job_id":       j.job_id,
            "scraper":      j.scraper,
            "status":       j.status,
            "date_from":    j.date_from,
            "date_to":      j.date_to,
            "started_at":   j.started_at.isoformat() if j.started_at  else None,
            "finished_at":  j.finished_at.isoformat() if j.finished_at else None,
            "record_count": j.record_count,
            "error_message": j.error_message,
        }
        for j in jobs
    ])


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8001, reload=True)

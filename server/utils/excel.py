import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

def _sanitize_cell(value):
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value

_HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_REJECT_FILL  = PatternFill("solid", fgColor="FFCCCC")


def styled_workbook(sheet_name: str, headers: list[str], rows: list[list]) -> io.BytesIO:
    """Build a styled openpyxl workbook and return it as an in-memory BytesIO stream."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = [_sanitize_cell(h) for h in headers]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[1].height = 30

    dec_idx = headers.index("Decision") if "Decision" in headers else -1

    for row_idx, row in enumerate(rows, start=2):
        sanitized_row = [_sanitize_cell(v) for v in row]
        ws.append(sanitized_row)
        if dec_idx != -1 and sanitized_row[dec_idx] == "REJECT":
            for cell in ws[row_idx]:
                cell.fill = _REJECT_FILL

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
